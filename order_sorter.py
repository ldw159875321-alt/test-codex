import argparse
import pathlib
import unicodedata
from typing import Iterable, Dict, List, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


COLUMN_ALIASES: Dict[str, List[str]] = {
    "order_id": ["order_id", "订单号", "注文番号", "orderno", "注文番号"],
    "customer_name": ["customer_name", "姓名", "氏名", "お名前", "名前"],
    "address": ["address", "住所", "收货地址"],
    "phone": ["phone", "電話", "phone_number", "tel", "電話番号", "电话"],
}

SEPARATOR_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


class OrderProcessingError(Exception):
    """Raised for predictable processing issues."""


def detect_column(df: pd.DataFrame, aliases: Iterable[str]) -> Optional[str]:
    for alias in aliases:
        if alias in df.columns:
            return alias
        # allow case-insensitive detection
        lower_matches = [c for c in df.columns if c.lower() == alias.lower()]
        if lower_matches:
            return lower_matches[0]
    return None


def standardize_columns(df: pd.DataFrame) -> pd.DataFrame:
    rename_map = {}
    for canonical, aliases in COLUMN_ALIASES.items():
        detected = detect_column(df, aliases)
        if detected:
            rename_map[detected] = canonical
    missing = [col for col in COLUMN_ALIASES if col not in rename_map.values()]
    if missing:
        raise OrderProcessingError(
            f"Missing required columns: {', '.join(missing)}. Available columns: {list(df.columns)}"
        )
    return df.rename(columns=rename_map)[list(COLUMN_ALIASES.keys())]


def normalize_text(value: str, *, strip_connectors: bool = False) -> str:
    if pd.isna(value):
        return ""
    normalized = unicodedata.normalize("NFKC", str(value))
    if strip_connectors:
        normalized = "".join(ch for ch in normalized if ch not in "-‐‑‒–—―ー－_ " and not ch.isspace())
    return normalized.strip()


def normalize_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["order_id"] = df["order_id"].map(normalize_text)
    df["customer_name"] = df["customer_name"].map(normalize_text)
    df["address"] = df["address"].map(lambda v: normalize_text(v, strip_connectors=True))
    df["phone"] = df["phone"].map(lambda v: normalize_text(v, strip_connectors=True))
    return df


def load_csv(path: pathlib.Path) -> pd.DataFrame:
    for encoding in ("shift_jis", "cp932", "utf-8-sig", "utf-8"):
        try:
            return pd.read_csv(path, encoding=encoding)
        except UnicodeDecodeError:
            continue
    raise OrderProcessingError(f"Unable to decode CSV file: {path}")


def load_excel(path: pathlib.Path) -> pd.DataFrame:
    try:
        return pd.read_excel(path)
    except Exception as exc:  # pragma: no cover - passthrough for clarity
        raise OrderProcessingError(f"Failed to read Excel file {path}: {exc}") from exc


def build_customer_key(df: pd.DataFrame) -> pd.Series:
    return (
        df["customer_name"].fillna("")
        + "|"
        + df["address"].fillna("")
        + "|"
        + df["phone"].fillna("")
    )


def write_with_separator(self_orders: pd.DataFrame, mixed_orders: pd.DataFrame, path: pathlib.Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        self_orders.to_excel(writer, index=False)
        startrow = len(self_orders) + 1 + 3
        mixed_orders.to_excel(writer, index=False, startrow=startrow)

    wb = load_workbook(path)
    ws = wb.active
    separator_start = len(self_orders) + 2  # account for header row
    for row_idx in range(separator_start, separator_start + 3):
        for col_idx in range(1, len(self_orders.columns) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is None:
                cell.value = ""
            cell.fill = SEPARATOR_FILL
    wb.save(path)


def process_orders(history_path: pathlib.Path, rms_path: pathlib.Path, boss_path: Optional[pathlib.Path], output_dir: pathlib.Path) -> None:
    history_df = standardize_columns(normalize_dataframe(load_excel(history_path)))
    rms_df = standardize_columns(normalize_dataframe(load_csv(rms_path)))
    boss_df = None
    if boss_path:
        boss_df = standardize_columns(normalize_dataframe(load_csv(boss_path)))

    # Step 2: de-duplication against history
    new_orders = rms_df[~rms_df["order_id"].isin(history_df["order_id"])]

    # Step 3: warehouse sorting
    if boss_df is None:
        rsl_orders = pd.DataFrame(columns=new_orders.columns)
        self_orders = new_orders
    else:
        boss_ids = set(boss_df["order_id"])
        rsl_orders = new_orders[new_orders["order_id"].isin(boss_ids)]
        self_orders = new_orders[~new_orders["order_id"].isin(boss_ids)]

    # Step 4: mixed order detection
    rsl_keys = set(build_customer_key(rsl_orders))
    self_orders = self_orders.copy()
    self_orders["customer_key"] = build_customer_key(self_orders)
    mixed_mask = self_orders["customer_key"].isin(rsl_keys)
    mixed_orders = self_orders[mixed_mask].drop(columns=["customer_key"])
    pure_self_orders = self_orders[~mixed_mask].drop(columns=["customer_key"])

    # Outputs
    output_dir.mkdir(parents=True, exist_ok=True)
    rsl_path = output_dir / "乐天仓发货单.xlsx"
    self_path = output_dir / "自发货单.xlsx"
    history_output_path = output_dir / "更新后的历史总表.xlsx"

    rsl_orders.to_excel(rsl_path, index=False)
    write_with_separator(pure_self_orders, mixed_orders, self_path)

    updated_history = pd.concat([history_df, new_orders], ignore_index=True)
    updated_history.to_excel(history_output_path, index=False)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Rakuten Order Sorter & De-duplicator")
    parser.add_argument("history", type=pathlib.Path, help="历史查重总表 (Excel)")
    parser.add_argument("rms", type=pathlib.Path, help="今日 RMS 订单 (CSV)")
    parser.add_argument("boss", nargs="?", type=pathlib.Path, help="今日 BOSS 订单 (CSV, 可选)")
    parser.add_argument("--output", type=pathlib.Path, default=pathlib.Path.cwd(), help="输出目录，默认当前路径")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    try:
        process_orders(args.history, args.rms, args.boss, args.output)
    except OrderProcessingError as exc:
        raise SystemExit(f"处理失败: {exc}")


if __name__ == "__main__":
    main()
